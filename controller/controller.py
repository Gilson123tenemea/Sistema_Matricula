# -*- coding: utf-8 -*-
"""
controller/controller.py
Lógica de negocio: autenticación, CRUD completo, reportes, PDF, Excel.
Roles: admin, docente, estudiante.
"""
import os
import random
import string
from utils.constants import NOTA_MINIMA_APROBACION, PERIODO, PERIODO_STR
from model.estudiante import Estudiante
from model.docente    import Docente
from model.carrera    import Carrera
from model.asignatura import Asignatura
from model.matricula  import Matricula
from model.usuario    import Usuario
from model.horario    import Horario, BloqueHorario


class Controller:
    def __init__(self, data):
        self.data = data
        self._usuario_activo: Usuario = None

    # ═══════ AUTENTICACIÓN ═══════════════════════
    def login(self, username: str, password: str) -> tuple:
        u = next((u for u in self.data.usuarios if u.username == username), None)
        if not u:
            return False, "Usuario no encontrado."
        if not u.verificar_password(password):
            return False, "Contraseña incorrecta."
        self._usuario_activo = u
        return True, u

    def logout(self):
        self._usuario_activo = None

    def es_admin(self) -> bool:
        return self._usuario_activo and self._usuario_activo.es_admin()

    def es_docente(self) -> bool:
        return self._usuario_activo and self._usuario_activo.es_docente()

    def cedula_activa(self) -> str:
        if self._usuario_activo:
            return self._usuario_activo.cedula_vinculada
        return None

    def rol_activo(self) -> str:
        return self._usuario_activo.rol if self._usuario_activo else ""

    @staticmethod
    def _generar_password(longitud=8) -> str:
        chars = string.ascii_letters + string.digits
        return ''.join(random.choices(chars, k=longitud))

    def crear_usuario(self, username, password, rol,
                       cedula_vinculada=None) -> tuple:
        if any(u.username == username for u in self.data.usuarios):
            return False, "Ese nombre de usuario ya existe.", None
        ph = Usuario.hash_password(password)
        u = Usuario(username, ph, rol, cedula_vinculada)
        self.data.usuarios.append(u)
        self.data.guardar()
        return True, f"Usuario '{username}' creado.", password

    def cambiar_password(self, username, old_pass, new_pass) -> tuple:
        u = next((u for u in self.data.usuarios if u.username == username), None)
        if not u:
            return False, "Usuario no encontrado."
        if not u.verificar_password(old_pass):
            return False, "Contraseña actual incorrecta."
        u.password_hash = Usuario.hash_password(new_pass)
        self.data.guardar()
        return True, "Contraseña actualizada."

    # Alias viejo
    def crear_usuario_estudiante(self, username, password, cedula_estudiante) -> tuple:
        return self.crear_usuario(username, password, "estudiante", cedula_estudiante)

    # ═══════ CARRERAS ════════════════════════════
    def registrar_carrera(self, codigo, nombre, num_ciclos=8, desc="") -> tuple:
        if any(c.codigo == codigo for c in self.data.carreras):
            return False, "Ya existe una carrera con ese código."
        if any(c.nombre.lower() == nombre.lower() for c in self.data.carreras):
            return False, "Ya existe una carrera con ese nombre."
        c = Carrera(codigo, nombre, num_ciclos, desc)
        self.data.carreras.append(c)
        self.data.guardar()
        return True, f"Carrera '{nombre}' registrada."

    def editar_carrera(self, codigo, nombre, num_ciclos=8, desc="") -> tuple:
        c = self._dic_carreras().get(codigo)
        if not c:
            return False, "Carrera no encontrada."
        c.nombre     = nombre
        c.num_ciclos = num_ciclos
        c.descripcion = desc
        self.data.guardar()
        return True, "Carrera actualizada."

    def eliminar_carrera(self, codigo) -> tuple:
        c = self._dic_carreras().get(codigo)
        if not c:
            return False, "Carrera no encontrada."
        # Verificar si hay estudiantes o asignaturas
        if any(e.carrera == codigo for e in self.data.estudiantes):
            return False, "No se puede eliminar: hay estudiantes en esta carrera."
        if any(a.carrera == codigo for a in self.data.asignaturas):
            return False, "No se puede eliminar: hay asignaturas en esta carrera."
        self.data.carreras.remove(c)
        self.data.guardar()
        return True, f"Carrera '{c.nombre}' eliminada."

    # ═══════ DOCENTES ═════════════════════════════
    def registrar_docente(self, cedula, nombres, titulo, email,
                           especialidad, telefono) -> tuple:
        from utils.validators import (validar_cedula, validar_email,
                                       validar_telefono, validar_nombre)
        if not validar_cedula(cedula):
            return False, "Cédula inválida (solo dígitos, 6-13 chars).", None
        if not validar_nombre(nombres):
            return False, "Nombre inválido (mínimo 3 caracteres, solo letras).", None
        if not validar_email(email):
            return False, "Correo electrónico inválido.", None
        if not validar_telefono(telefono):
            return False, "Teléfono inválido (7-15 dígitos).", None
        if cedula in self._dic_docentes():
            return False, "Ya existe un docente con esa cédula.", None
        d = Docente(cedula, nombres, titulo, email, especialidad, telefono)
        self.data.docentes.append(d)
        # Generar credenciales
        username = "doc_" + cedula[:6]
        # Evitar duplicados
        base = username
        i = 1
        while any(u.username == username for u in self.data.usuarios):
            username = f"{base}_{i}"; i += 1
        password = self._generar_password()
        self.crear_usuario(username, password, "docente", cedula)
        return True, f"Docente '{nombres}' registrado.", (username, password)

    def editar_docente(self, cedula, nombres, titulo, email,
                        especialidad, telefono) -> tuple:
        from utils.validators import validar_email, validar_telefono
        if not validar_email(email):
            return False, "Correo electrónico inválido."
        if not validar_telefono(telefono):
            return False, "Teléfono inválido."
        d = self._dic_docentes().get(cedula)
        if not d:
            return False, "Docente no encontrado."
        d.nombres = nombres; d.titulo = titulo; d.email = email
        d.especialidad = especialidad; d.telefono = telefono
        self.data.guardar()
        return True, "Docente actualizado correctamente."

    def eliminar_docente(self, cedula) -> tuple:
        d = self._dic_docentes().get(cedula)
        if not d:
            return False, "Docente no encontrado."
        # Desvincular asignaturas
        for a in self.data.asignaturas:
            if a.docente == cedula:
                a.docente = ""
        # Eliminar usuario asociado
        self.data.usuarios = [u for u in self.data.usuarios
                               if u.cedula_vinculada != cedula]
        self.data.docentes.remove(d)
        self.data.guardar()
        return True, f"Docente '{d.nombres}' eliminado."

    def asignar_docente_a_materia(self, cedula_docente, codigo_asig) -> tuple:
        a = self._dic_asignaturas().get(codigo_asig)
        d = self._dic_docentes().get(cedula_docente)
        if not a:
            return False, "Asignatura no encontrada."
        if not d:
            return False, "Docente no encontrado."
        a.docente = cedula_docente
        self.data.guardar()
        return True, f"Docente '{d.nombres}' asignado a '{a.nombre}'."

    # ═══════ HORARIOS ══════════════════════════════
    def registrar_horario(self, bloques_data: list, aula: str = "") -> tuple:
        """
        bloques_data: list of (dia, hora_inicio, hora_fin)
        Crea un horario con múltiples bloques y lo guarda.
        """
        from model.horario import Horario, BloqueHorario
        from utils.constants import DIAS_SEMANA
        if not bloques_data:
            return False, "Debes agregar al menos un bloque horario.", None

        bloques = []
        for dia, hi, hf in bloques_data:
            if dia not in DIAS_SEMANA:
                return False, f"Día inválido: {dia}", None
            if hi >= hf:
                return False, f"Hora inicio debe ser menor que hora fin en {dia}.", None
            bloques.append(BloqueHorario(dia, hi, hf))

        # Verificar que los bloques del mismo horario no se solapen entre sí
        for i, b1 in enumerate(bloques):
            for j, b2 in enumerate(bloques):
                if i != j and b1.solapado(b2):
                    return False, f"Los bloques del horario se solapan: {b1} y {b2}.", None

        h = Horario(None, bloques, aula or "Por asignar")
        nuevo_id = self.data.insertar_horario(h.bloques_str(), h.aula)
        h.id = nuevo_id
        self.data.horarios.append(h)
        return True, f"Horario registrado: {h.resumen()}", h

    def eliminar_horario(self, horario_id: int) -> tuple:
        h = next((x for x in self.data.horarios if x.id == horario_id), None)
        if not h:
            return False, "Horario no encontrado."
        for a in self.data.asignaturas:
            if a.horario_id == horario_id:
                a.horario_id = None
        self.data.horarios.remove(h)
        self.data.guardar()
        return True, "Horario eliminado."

    def horarios_disponibles(self) -> list:
        return self.data.horarios

    def calcular_bloques_sugeridos(self, horas_semanales: int) -> list:
        """
        Calcula la distribución óptima de bloques según horas semanales.
        Retorna lista de (num_dias, horas_por_bloque) con los mejores patrones.
        """
        sugerencias = []
        if horas_semanales <= 2:
            sugerencias.append((1, horas_semanales))        # Todo en 1 día
        elif horas_semanales == 3:
            sugerencias.append((3, 1))                       # 1h x 3 días
            sugerencias.append((1, 3))                       # 3h x 1 día
        elif horas_semanales == 4:
            sugerencias.append((2, 2))                       # 2h x 2 días ← ideal
            sugerencias.append((4, 1))                       # 1h x 4 días
        elif horas_semanales == 5:
            sugerencias.append((2, 2))                       # 2+3 aprox
            sugerencias.append((5, 1))
        elif horas_semanales == 6:
            sugerencias.append((3, 2))                       # 2h x 3 días ← ideal
            sugerencias.append((2, 3))                       # 3h x 2 días
        elif horas_semanales == 8:
            sugerencias.append((4, 2))                       # 2h x 4 días
            sugerencias.append((2, 4))                       # 4h x 2 días
        else:
            # Genérico
            for n in [2, 3, 4]:
                if horas_semanales % n == 0:
                    sugerencias.append((n, horas_semanales // n))
            sugerencias.append((1, horas_semanales))
        return sugerencias

    def verificar_conflicto_docente(self, cedula_docente: str,
                                     nuevo_horario,
                                     excluir_asig: str = None) -> tuple:
        """
        Verifica si el docente tiene conflicto de horario con sus otras materias.
        Retorna (tiene_conflicto: bool, detalle: str)
        """
        if not cedula_docente or not nuevo_horario:
            return False, ""
        for a in self.data.asignaturas:
            if a.docente != cedula_docente: continue
            if excluir_asig and a.codigo == excluir_asig: continue
            if not a.horario_id: continue
            h_existente = self.data.horario_by_id(a.horario_id)
            if h_existente and nuevo_horario.tiene_conflicto_con(h_existente):
                return True, (f"El docente ya tiene la materia '{a.nombre}' "
                              f"en el horario: {h_existente.resumen()}")
        return False, ""

    def asignar_docente_con_horario(self, cedula_docente: str,
                                     codigo_asig: str,
                                     bloques_data: list,
                                     aula: str = "") -> tuple:
        """
        Flujo integrado: asigna docente + crea/actualiza horario + verifica conflictos.
        bloques_data: list of (dia, hora_inicio, hora_fin)
        """
        from model.horario import Horario, BloqueHorario
        a = self._dic_asignaturas().get(codigo_asig)
        d = self._dic_docentes().get(cedula_docente)
        if not a: return False, "Asignatura no encontrada."
        if not d: return False, "Docente no encontrado."

        # Construir el nuevo horario (sin guardar aún)
        bloques = []
        for dia, hi, hf in bloques_data:
            bloques.append(BloqueHorario(dia, hi, hf))
        nuevo_h = Horario(None, bloques, aula or "Por asignar")

        # Verificar conflicto
        conflicto, detalle = self.verificar_conflicto_docente(
            cedula_docente, nuevo_h, excluir_asig=codigo_asig)
        if conflicto:
            return False, f"⚠ Conflicto de horario: {detalle}"

        # Eliminar horario anterior de esta asignatura si existe
        if a.horario_id:
            h_viejo = self.data.horario_by_id(a.horario_id)
            if h_viejo:
                self.data.horarios.remove(h_viejo)

        # Guardar nuevo horario
        nuevo_id = self.data.insertar_horario(nuevo_h.bloques_str(), nuevo_h.aula)
        nuevo_h.id = nuevo_id
        self.data.horarios.append(nuevo_h)

        # Asignar docente y horario
        a.docente    = cedula_docente
        a.horario_id = nuevo_id
        self.data.guardar()

        horas_reales = nuevo_h.horas_semanales()
        return True, (f"✅ Docente '{d.nombres}' asignado a '{a.nombre}'.\n"
                      f"Horario: {nuevo_h.resumen()}\n"
                      f"Aula: {nuevo_h.aula}\n"
                      f"Total: {horas_reales}h semanales")

    # ═══════ ESTUDIANTES ══════════════════════════
    def registrar_estudiante(self, cedula, nombre, edad,
                              carrera, email, semestre=1) -> tuple:
        from utils.validators import validar_cedula, validar_email, validar_nombre
        if not validar_cedula(cedula):
            return False, "Cédula inválida.", None
        if not validar_nombre(nombre):
            return False, "Nombre inválido.", None
        if not validar_email(email):
            return False, "Email inválido.", None
        if cedula in self._dic_estudiantes():
            return False, "Ya existe un estudiante con esa cédula.", None
        if not self._dic_carreras().get(carrera):
            return False, "Carrera no encontrada. Regístrala primero.", None
        e = Estudiante(cedula, nombre, edad, carrera, email, semestre)
        self.data.estudiantes.append(e)
        # Generar credenciales
        username = "est_" + cedula[:6]
        base = username; i = 1
        while any(u.username == username for u in self.data.usuarios):
            username = f"{base}_{i}"; i += 1
        password = self._generar_password()
        self.crear_usuario(username, password, "estudiante", cedula)
        return True, f"Estudiante '{nombre}' registrado.", (username, password)

    def editar_estudiante(self, cedula, nombre, edad,
                          carrera, email, semestre=1) -> tuple:
        from utils.validators import validar_email
        e = self._dic_estudiantes().get(cedula)
        if not e:
            return False, "Estudiante no encontrado."
        if not validar_email(email):
            return False, "Email inválido."
        e.nombre = nombre; e.edad = edad; e.carrera = carrera
        e.email = email; e.semestre = semestre
        self.data.guardar()
        return True, "Estudiante actualizado correctamente."

    def eliminar_estudiante(self, cedula) -> tuple:
        e = self._dic_estudiantes().get(cedula)
        if not e:
            return False, "Estudiante no encontrado.", 0
        mats_rel = [m for m in self.data.matriculas if m.estudiante.cedula == cedula]
        for m in mats_rel:
            self.data.matriculas.remove(m)
        self.data.estudiantes.remove(e)
        self.data.usuarios = [u for u in self.data.usuarios
                               if u.cedula_vinculada != cedula]
        self.data.guardar()
        return True, f"Estudiante '{e.nombre}' eliminado.", len(mats_rel)

    def busqueda_avanzada(self, carrera="", semestre="", estado="",
                          docente="", materia="") -> list:
        return self.data.query_matriculas_join(carrera, semestre, estado,
                                               docente, materia)

    # ═══════ ASIGNATURAS ══════════════════════════
    def registrar_asignatura(self, codigo, nombre, docente, creditos,
                              cupo_maximo=30, semestre=1, prerequisito=None,
                              carrera="", horario_id=None, horas_semanales=None) -> tuple:
        from utils.validators import validar_creditos, validar_codigo
        if not validar_codigo(codigo):
            return False, "Código inválido (3-20 chars alfanuméricos)."
        ok, cr = validar_creditos(str(creditos))
        if not ok:
            return False, "Créditos inválidos (1-10)."
        if codigo in self._dic_asignaturas():
            return False, "Ya existe una asignatura con ese código."
        if prerequisito and prerequisito not in self._dic_asignaturas():
            return False, "El prerequisito indicado no existe."
        hs = horas_semanales or cr
        a = Asignatura(codigo, nombre, docente, cr, cupo_maximo,
                        semestre, prerequisito or None, carrera, horario_id, hs)
        self.data.asignaturas.append(a)
        self.data.guardar()
        return True, f"Asignatura '{nombre}' registrada."

    def editar_asignatura(self, codigo, nombre, docente, creditos,
                          cupo_maximo=30, semestre=1, prerequisito=None,
                          carrera="", horario_id=None, horas_semanales=None) -> tuple:
        a = self._dic_asignaturas().get(codigo)
        if not a:
            return False, "Asignatura no encontrada."
        a.nombre = nombre; a.docente = docente; a.creditos = creditos
        a.cupo_maximo = cupo_maximo; a.semestre = semestre
        a.prerequisito = prerequisito or None
        a.carrera = carrera; a.horario_id = horario_id
        if horas_semanales: a.horas_semanales = horas_semanales
        self.data.guardar()
        return True, "Asignatura actualizada."

    def eliminar_asignatura(self, codigo) -> tuple:
        a = self._dic_asignaturas().get(codigo)
        if not a:
            return False, "Asignatura no encontrada.", 0
        mats_rel = [m for m in self.data.matriculas if m.asignatura.codigo == codigo]
        for m in mats_rel:
            self.data.matriculas.remove(m)
        self.data.asignaturas.remove(a)
        self.data.guardar()
        return True, f"Asignatura '{a.nombre}' eliminada.", len(mats_rel)

    def cupos_disponibles(self, codigo: str) -> int:
        a = self._dic_asignaturas().get(codigo)
        if not a:
            return 0
        usados = self.data.query_cupos_usados(codigo)
        return max(0, a.cupo_maximo - usados)

    def materias_por_carrera_ciclo(self, carrera: str, ciclo: int) -> list:
        return [a for a in self.data.asignaturas
                if a.carrera == carrera and a.semestre == ciclo]

    # ═══════ MATRÍCULAS ═══════════════════════════
    def registrar_matricula(self, cedula, codigo, nota=0.0) -> tuple:
        e = self._dic_estudiantes().get(cedula)
        a = self._dic_asignaturas().get(codigo)
        if not e or not a:
            return False, "Estudiante o asignatura no encontrados."
        if self.cupos_disponibles(codigo) <= 0:
            return False, f"No hay cupos disponibles en '{a.nombre}'."
        if a.prerequisito:
            aprobado = any(
                m.estudiante.cedula == cedula
                and m.asignatura.codigo == a.prerequisito
                and m.nota >= NOTA_MINIMA_APROBACION
                for m in self.data.matriculas
            )
            if not aprobado:
                pre = self._dic_asignaturas().get(a.prerequisito)
                pnombre = pre.nombre if pre else a.prerequisito
                return False, f"Prerequisito no aprobado: '{pnombre}'."
        ya = any(m.estudiante.cedula == cedula and m.asignatura.codigo == codigo
                 for m in self.data.matriculas)
        if ya:
            return False, "El estudiante ya está matriculado en esa asignatura."
        m = Matricula(e, a, nota, PERIODO_STR)
        self.data.matriculas.append(m)
        self.data.guardar_historial_row(cedula, codigo, a.nombre,
                                         nota, PERIODO_STR, a.creditos)
        self.data.guardar()
        return True, (f"Matrícula registrada.\n\n"
                      f"Estudiante : {e.nombre}\n"
                      f"Asignatura : {a.nombre}\n"
                      f"Cupos rest.: {self.cupos_disponibles(codigo)}")

    def matricular_multiples(self, cedula: str, codigos: list) -> tuple:
        """Matricula al estudiante en múltiples materias a la vez."""
        errores = []
        ok_count = 0
        for cod in codigos:
            ok, msg = self.registrar_matricula(cedula, cod)
            if ok:
                ok_count += 1
            else:
                errores.append(f"• {cod}: {msg}")
        if errores:
            return ok_count > 0, (f"{ok_count} materias matriculadas.\n"
                                   f"Errores:\n" + "\n".join(errores))
        return True, f"{ok_count} materias matriculadas correctamente."

    def editar_nota(self, cedula, codigo, nueva_nota) -> tuple:
        for m in self.data.matriculas:
            if m.estudiante.cedula == cedula and m.asignatura.codigo == codigo:
                m.nota = nueva_nota
                self.data.guardar_historial_row(
                    cedula, codigo, m.asignatura.nombre,
                    nueva_nota, m.periodo, m.asignatura.creditos)
                self.data.guardar()
                return True, "Nota actualizada correctamente."
        return False, "Matrícula no encontrada."

    def estudiantes_disponibles(self, codigo_asig: str) -> list:
        ya = {m.estudiante.cedula for m in self.data.matriculas
              if m.asignatura.codigo == codigo_asig}
        return [e for e in self.data.estudiantes if e.cedula not in ya]

    # ═══════ HISTORIAL ════════════════════════════
    def historial_estudiante(self, cedula: str) -> list:
        return self.data.query_historial(cedula)

    # ═══════ REPORTES TEXTO ═══════════════════════
    def reporte_promedio(self, cedula, incluir_ponderado=True) -> str:
        e = self._dic_estudiantes().get(cedula)
        if not e:
            return "Estudiante no encontrado."
        mats = [m for m in self.data.matriculas if m.estudiante.cedula == cedula]
        if not mats:
            return f"El estudiante '{e.nombre}' no tiene notas registradas."
        prom_simple = sum(m.nota for m in mats) / len(mats)
        total_cred  = sum(m.asignatura.creditos for m in mats)
        prom_pond   = (sum(m.nota * m.asignatura.creditos for m in mats)
                       / total_cred if total_cred else 0)
        lin = ["="*62, "  PROMEDIO DEL ESTUDIANTE", "="*62,
               f"  Estudiante : {e.nombre}", f"  Cédula     : {e.cedula}",
               f"  Carrera    : {e.carrera}", f"  Semestre   : {e.semestre}",
               "-"*62,
               f"  {'Asignatura':<26} {'Cred.':<7} {'Nota':<8} Estado",
               "-"*62]
        for m in mats:
            lin.append(f"  {m.asignatura.nombre:<26} {m.asignatura.creditos:<7}"
                       f" {m.nota:<8.2f} {m.estado()}")
        lin += ["-"*62, f"  Promedio simple    : {prom_simple:.2f}"]
        if incluir_ponderado:
            lin.append(f"  Promedio ponderado : {prom_pond:.2f}  (por créditos)")
        lin += [f"  Materias cursadas  : {len(mats)}", "="*62]
        return "\n".join(lin)

    def reporte_estados(self) -> str:
        if not self.data.matriculas:
            return "No hay matrículas registradas."
        ap  = [m for m in self.data.matriculas if m.nota >= NOTA_MINIMA_APROBACION]
        rep = [m for m in self.data.matriculas if m.nota < NOTA_MINIMA_APROBACION]
        tasa = len(ap) / len(self.data.matriculas) * 100
        lin = ["="*62, "  REPORTE  APROBADOS / REPROBADOS", "="*62,
               f"\n  Aprobados ({len(ap)})", "  " + "-"*58]
        if ap:
            lin.append(f"  {'Estudiante':<26} {'Asignatura':<26} Nota")
            for m in ap:
                lin.append(f"  {m.estudiante.nombre:<26}"
                            f" {m.asignatura.nombre:<26} {m.nota:.2f}")
        else:
            lin.append("  Ninguno.")
        lin += [f"\n  Reprobados ({len(rep)})", "  " + "-"*58]
        if rep:
            lin.append(f"  {'Estudiante':<26} {'Asignatura':<26} Nota")
            for m in rep:
                lin.append(f"  {m.estudiante.nombre:<26}"
                            f" {m.asignatura.nombre:<26} {m.nota:.2f}")
        else:
            lin.append("  Ninguno.")
        lin += [f"\n  Tasa de aprobación : {tasa:.1f}%", "="*62]
        return "\n".join(lin)

    def reporte_ranking(self) -> str:
        promedios = []
        for e in self.data.estudiantes:
            notas = [m.nota for m in self.data.matriculas
                     if m.estudiante.cedula == e.cedula]
            if notas:
                promedios.append((e, sum(notas)/len(notas), len(notas)))
        if not promedios:
            return "No hay notas registradas para generar el ranking."
        promedios.sort(key=lambda x: x[1], reverse=True)
        lin = ["="*72, "  RANKING DE ESTUDIANTES", "="*72,
               f"  {'Pos':<5} {'Nombre':<26} {'Carrera':<21} {'Prom.':<8} Materias",
               "  " + "-"*65]
        for i, (e, prom, n) in enumerate(promedios, 1):
            lin.append(f"  {i:<5} {e.nombre:<26} {e.carrera:<21}"
                        f" {prom:<8.2f} {n}")
        lin.append("="*72)
        return "\n".join(lin)

    def reporte_estadisticas(self) -> str:
        stats = self.data.estadisticas_sql()
        nombres_carreras = {c.codigo: c.nombre for c in self.data.carreras}
        lin = ["="*52, "  ESTADÍSTICAS GENERALES", "="*52,
               f"  Período            : {PERIODO[0]}-{PERIODO[1]}",
               f"  Total estudiantes  : {len(self.data.estudiantes)}",
               f"  Total asignaturas  : {len(self.data.asignaturas)}",
               f"  Total matrículas   : {len(self.data.matriculas)}",
               f"  Carreras activas   : {len(self.data.carreras)}",
               f"  Docentes           : {len(self.data.docentes)}"]
        if stats and stats["total_matriculas"]:
            tasa = (stats["aprobados"] / stats["total_matriculas"] * 100
                    if stats["total_matriculas"] else 0)
            lin += ["  " + "-"*48,
                    f"  Nota máxima        : {stats['max_nota']:.2f}",
                    f"  Nota mínima        : {stats['min_nota']:.2f}",
                    f"  Promedio general   : {stats['promedio']:.2f}",
                    f"  Tasa de aprobación : {tasa:.1f}%",
                    f"  Aprobados          : {stats['aprobados']}",
                    f"  Reprobados         : {stats['reprobados']}"]
        if self.data.carreras:
            lin.append("\n  Carreras registradas:")
            for c in self.data.carreras:
                lin.append(f"    - {c.nombre}")
        lin.append("="*52)
        return "\n".join(lin)

    # ═══════ PDF MATRÍCULA ════════════════════════
    def generar_pdf_matricula(self, cedula: str, ruta_salida: str) -> tuple:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import (SimpleDocTemplate, Table,
                                             TableStyle, Paragraph,
                                             Spacer, HRFlowable)
            from reportlab.lib.units import cm
            import datetime
        except ImportError:
            return False, "reportlab no instalado. Ejecute: pip install reportlab"

        e = self._dic_estudiantes().get(cedula)
        if not e:
            return False, "Estudiante no encontrado."
        mats = [m for m in self.data.matriculas if m.estudiante.cedula == cedula]
        nombre_carrera = next((c.nombre for c in self.data.carreras
                                if c.codigo == e.carrera), e.carrera)
        # Obtener nombre del usuario
        u = next((u for u in self.data.usuarios
                  if u.cedula_vinculada == cedula and u.rol == "estudiante"), None)

        doc = SimpleDocTemplate(ruta_salida, pagesize=A4,
                                 rightMargin=2*cm, leftMargin=2*cm,
                                 topMargin=2.5*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        azul   = colors.HexColor("#1f6feb")
        oscuro = colors.HexColor("#111820")
        gris   = colors.HexColor("#6e7681")
        verde  = colors.HexColor("#2ea043")
        rojo   = colors.HexColor("#da3633")

        title_style = ParagraphStyle("Title2", parent=styles["Title"],
                                      textColor=azul, fontSize=20, spaceAfter=4)
        sub_style   = ParagraphStyle("Sub", parent=styles["Normal"],
                                      textColor=gris, fontSize=10)
        bold_style  = ParagraphStyle("Bold", parent=styles["Normal"],
                                      textColor=oscuro, fontSize=11,
                                      fontName="Helvetica-Bold")
        normal = styles["Normal"]

        story = []
        story.append(Paragraph("SISTEMA ACADÉMICO UNIVERSITARIO", title_style))
        story.append(Paragraph(f"Comprobante de Matrícula — Período {PERIODO_STR}", sub_style))
        story.append(HRFlowable(width="100%", thickness=1, color=azul, spaceAfter=12))
        story.append(Spacer(1, 0.3*cm))

        info = [
            ["Nombre:",  e.nombre,       "Cédula:",   e.cedula],
            ["Carrera:", nombre_carrera, "Ciclo:",    str(e.semestre)],
            ["Email:",   e.email,        "Usuario:",  u.username if u else "—"],
        ]
        t_info = Table(info, colWidths=[3*cm, 6.5*cm, 3*cm, 4*cm])
        t_info.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (0, -1), azul),
            ("TEXTCOLOR", (2, 0), (2, -1), azul),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t_info)
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=gris, spaceAfter=8))
        story.append(Paragraph("<b>Asignaturas Matriculadas</b>", normal))
        story.append(Spacer(1, 0.2*cm))

        # Obtener nombre docente
        dic_doc = {d.cedula: d.nombres for d in self.data.docentes}
        encabezado = ["#", "Código", "Asignatura", "Docente", "Créditos", "Estado"]
        filas = [encabezado]
        for i, m in enumerate(mats, 1):
            doc_nombre = dic_doc.get(m.asignatura.docente, m.asignatura.docente)
            filas.append([str(i), m.asignatura.codigo, m.asignatura.nombre,
                          doc_nombre, str(m.asignatura.creditos), "Matriculado"])

        total_cred = sum(m.asignatura.creditos for m in mats)
        filas.append(["", "", "", "TOTAL CRÉDITOS", str(total_cred), ""])

        t_mats = Table(filas, colWidths=[0.8*cm, 2*cm, 5.5*cm, 3.5*cm, 1.8*cm, 2.4*cm])
        ts = TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), azul),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 9),
            ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",        (2, 1), (3, -1), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2),
             [colors.white, colors.HexColor("#f0f4ff")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cccccc")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8f0fe")),
        ])
        t_mats.setStyle(ts)
        story.append(t_mats)
        story.append(Spacer(1, 0.8*cm))
        story.append(Paragraph(
            f"Documento generado el {datetime.date.today().strftime('%d/%m/%Y')}  "
            f"| Sistema Académico Universitario  |  Período {PERIODO_STR}", sub_style))
        doc.build(story)
        return True, f"PDF generado: {ruta_salida}"

    # ═══════ EXCEL DOCENTE ════════════════════════
    def exportar_excel_curso(self, codigo_asig: str, ruta_salida: str) -> tuple:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return False, "openpyxl no instalado. Ejecute: pip install openpyxl"

        a = self._dic_asignaturas().get(codigo_asig)
        if not a:
            return False, "Asignatura no encontrada."
        mats = [m for m in self.data.matriculas if m.asignatura.codigo == codigo_asig]
        dic_doc = {d.cedula: d.nombres for d in self.data.docentes}
        nombre_carrera = next((c.nombre for c in self.data.carreras
                                if c.codigo == a.carrera), a.carrera)
        doc_nombre = dic_doc.get(a.docente, a.docente)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista de Estudiantes"

        azul = "1F6FEB"; verde = "2EA043"; rojo = "DA3633"; gris = "F6F8FA"
        hdr_fill = PatternFill("solid", fgColor=azul)
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        bold     = Font(bold=True, size=11)
        thin     = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"))
        center = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:G1")
        ws["A1"] = f"Lista de Estudiantes — {a.nombre} ({a.codigo})"
        ws["A1"].font = Font(bold=True, size=14, color=azul)
        ws["A1"].alignment = center
        ws.merge_cells("A2:G2")
        ws["A2"] = (f"Docente: {doc_nombre}  |  Carrera: {nombre_carrera}  "
                    f"|  Créditos: {a.creditos}  |  Período: {PERIODO_STR}"
                    f"|  Cupo: {a.cupo_maximo}")
        ws["A2"].font = Font(size=10, color="6E7681")
        ws["A2"].alignment = center

        cols = ["#", "Cédula", "Nombre", "Carrera", "Ciclo", "Nota", "Estado"]
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=4, column=ci, value=col)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = thin

        for ri, m in enumerate(mats, 1):
            row_data = [ri, m.estudiante.cedula, m.estudiante.nombre,
                        m.estudiante.carrera, m.estudiante.semestre,
                        round(m.nota, 2), m.estado()]
            fill_row = PatternFill("solid", fgColor=(gris if ri % 2 == 0 else "FFFFFF"))
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=4+ri, column=ci, value=val)
                cell.border = thin; cell.alignment = center; cell.fill = fill_row
                if ci == 7:
                    cell.font = Font(bold=True, color=(verde if val == "Aprobado" else rojo))

        row_res = 4 + len(mats) + 2
        ap  = sum(1 for m in mats if m.estado() == "Aprobado")
        rep = len(mats) - ap
        prom = (sum(m.nota for m in mats) / len(mats) if mats else 0)
        for i, (label, val) in enumerate([
            ("Total matriculados:", len(mats)), ("Aprobados:", ap),
            ("Reprobados:", rep), ("Promedio del curso:", round(prom, 2))]):
            ws.cell(row=row_res+i, column=1, value=label).font = bold
            ws.cell(row=row_res+i, column=2, value=val)

        anchos = [5, 14, 28, 22, 7, 8, 12]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

        wb.save(ruta_salida)
        return True, f"Excel guardado: {ruta_salida}"

    # ═══════ DATOS PARA GRÁFICOS ══════════════════
    def datos_grafico_aprobados(self) -> dict:
        total = len(self.data.matriculas)
        if not total:
            return {"aprobados": 0, "reprobados": 0, "total": 0}
        ap  = sum(1 for m in self.data.matriculas if m.nota >= NOTA_MINIMA_APROBACION)
        return {"aprobados": ap, "reprobados": total - ap, "total": total}

    def datos_grafico_por_carrera(self) -> dict:
        por_carrera = {}
        for m in self.data.matriculas:
            c = m.estudiante.carrera
            por_carrera.setdefault(c, []).append(m.nota)
        return {c: sum(ns)/len(ns) for c, ns in por_carrera.items()}

    def stats_docente(self, cedula_docente: str) -> list:
        return self.data.stats_docente(cedula_docente)

    # ═══════ HELPERS ══════════════════════════════
    def _dic_estudiantes(self) -> dict:
        return {e.cedula: e for e in self.data.estudiantes}

    def _dic_asignaturas(self) -> dict:
        return {a.codigo: a for a in self.data.asignaturas}

    def _dic_docentes(self) -> dict:
        return {d.cedula: d for d in self.data.docentes}

    def _dic_carreras(self) -> dict:
        return {c.codigo: c for c in self.data.carreras}
